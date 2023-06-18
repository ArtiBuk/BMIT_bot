from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import User, Projects, Report
from rest_framework.serializers import ModelSerializer
from django.db.models import Sum
from django.shortcuts import render
from datetime import datetime, timedelta
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

class UserSerializer(ModelSerializer):
    class Meta:
        model = User
        fields = [
            "username",
            "first_name",
            "last_name",
            "birthday",
            "tg_id",
            "id",
        ]


class ProjectSerializer(ModelSerializer):
    users = UserSerializer(many=True)

    class Meta:
        model = Projects
        fields = ["id", "name", "short_description", "users"]


class ReportSerializer(ModelSerializer):
    class Meta:
        model = Report
        fields = ["date", "hours", "user", "project", "text_report"]


@api_view(["POST"])
def create_user(request):
    serializer = UserSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["POST"])
def create_report(request):
    serializer = ReportSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["GET"])
def view_users(request):
    users = User.objects.all()
    serializer = UserSerializer(users, many=True)
    return Response(serializer.data, status=status.HTTP_200_OK)


@api_view(["GET"])
def view_projects(request):
    projects = Projects.objects.all()
    serializer = ProjectSerializer(projects, many=True)
    return Response(serializer.data, status=status.HTTP_200_OK)


@api_view(["GET"])
def view_report(request):
    # Получение параметров запроса
    project_name = request.query_params.get("project_name")
    hours = request.query_params.get("hours")
    date = request.query_params.get("date")
    user_id = request.query_params.get("user_id")

    # Фильтрация данных отчета
    reports = Report.objects.all()

    if project_name:
        reports = reports.filter(project__name=project_name)

    if hours:
        reports = reports.filter(hours=hours)

    if date:
        reports = reports.filter(date=date)

    if user_id:
        reports = reports.filter(user_id=user_id)

    serializer = ReportSerializer(reports, many=True)
    return Response(serializer.data, status=status.HTTP_200_OK)

@login_required
def report_view(request):
    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        start_date = datetime.strptime(start_date, '%d.%m.%Y').date()
        end_date = datetime.strptime(end_date, '%d.%m.%Y').date()

        # Получение параметров фильтрации
        user_id = request.POST.get('user_id')
        project_id = request.POST.get('project_id')

        # Получение всех пользователей и проектов
        users = User.objects.all()
        projects = Projects.objects.all()

        # Создание заголовка таблицы
        table_header = ['Проект', 'Сотрудник']
        current_date = start_date
        while current_date <= end_date:
            table_header.append(current_date.strftime('%d.%m'))
            current_date += timedelta(days=1)

        # Создание таблицы с данными
        table_data = []
        for project in projects:
            for user in project.users.all():
                if user_id and user_id != str(user.id):
                    continue
                if project_id and project_id != str(project.id):
                    continue
                row = [project.name, user.get_full_name()]
                current_date = start_date
                while current_date <= end_date:
                    total_hours = Report.objects.filter(
                        date=current_date,
                        user=user,
                        project=project
                    ).aggregate(Sum('hours')).get('hours__sum')
                    row.append(total_hours or '')
                    current_date += timedelta(days=1)
                table_data.append(row)

        context = {
            'table_header': table_header,
            'table_data': table_data,
            'users': users,
            'projects': projects
        }


        if 'export' in request.POST:
            # Создание Excel-отчета
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=report.xlsx'

            workbook = Workbook()
            worksheet = workbook.active

            # Стилизация заголовка
            header_fill = PatternFill(start_color="FFD6DCE5", end_color="FFD6DCE5", fill_type="solid")
            header_font = Font(bold=True)
            header_border = Border(bottom=Side(border_style="thin"))
            header_alignment = Alignment(horizontal="center", vertical="center")

            for index, header in enumerate(table_header, start=1):
                cell = worksheet.cell(row=1, column=index, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = header_border
                cell.alignment = header_alignment

            # Стилизация данных
            data_fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
            weekend_fill = PatternFill(start_color="FFC0C0C0", end_color="FFC0C0C0", fill_type="solid")
            user_fill = PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid")
            project_fill = PatternFill(start_color="FF92D050", end_color="FF92D050", fill_type="solid")
            border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            data_alignment = Alignment(horizontal="center", vertical="center")

            for row_index, row_data in enumerate(table_data, start=2):
                for col_index, value in enumerate(row_data, start=1):
                    cell = worksheet.cell(row=row_index, column=col_index, value=value)
                    cell.fill = data_fill
                    cell.border = border

                    # Стилизация столбцов Проект и Сотрудник
                    if col_index == 1:
                        cell.fill = project_fill
                    elif col_index == 2:
                        cell.fill = user_fill

                    # Стилизация столбцов Дат и Количество часов
                    if col_index > 2:
                        current_date = start_date + timedelta(days=col_index - 3)
                        if current_date.weekday() >= 5:  # Проверка на выходные дни
                            cell.fill = weekend_fill

                    cell.alignment = data_alignment

            workbook.save(response)
            return response

        return render(request, 'report.html', context)
    else:
        users = User.objects.all()
        projects = Projects.objects.all()
        context = {
            'users': users,
            'projects': projects
        }
        return render(request, 'report.html', context)
    


